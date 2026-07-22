import openpyxl
import json
import os

def compile_seo_data():
    excel_path = r"c:\Users\amity\OneDrive\Desktop\saksi-forge\SteelManufacturer_Full_SEO_Optimization.xlsx"
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        return

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # 1. Load Keyword Map
    keyword_map = {}
    keyword_sheet = wb['7. Keyword Map']
    kw_headers = [c.value for c in keyword_sheet[1]]
    for row in list(keyword_sheet.iter_rows(values_only=True))[1:]:
        if row[0]:
            url = row[0].strip()
            keyword_map[url] = {
                'Primary Keyword': row[1],
                'Secondary Keywords': row[2],
                'Long-tail Keywords': row[3]
            }

    # 2. Load FAQ Bank
    faq_map = {}
    faq_sheet = wb['6. FAQ Bank (AEO)']
    faq_headers = [c.value for c in faq_sheet[1]]
    for row in list(faq_sheet.iter_rows(values_only=True))[1:]:
        if row[0]:
            url = row[0].strip()
            if url not in faq_map:
                faq_map[url] = []
            q = row[3]
            a = row[4]
            if q and a:
                faq_map[url].append({'q': str(q).strip(), 'a': str(a).strip()})

    # 3. Load Product Page Content
    product_content_map = {}
    content_sheet = wb['4. Product Page Content']
    content_headers = [c.value for c in content_sheet[1]]
    for row in list(content_sheet.iter_rows(values_only=True))[1:]:
        if row[0]:
            url = row[0].strip()
            product_content_map[url] = {
                'Product Name': row[1],
                'Priority': row[2],
                'TL;DR': row[3],
                'H2 Outline': row[4],
                'Opening Copy': row[5],
                'Specs Table Raw': row[6],
                'Internal Links Raw': row[7],
                'Image Alt Texts': row[8],
                'Word Count Target': row[9]
            }

    # 4. Load Page Inventory
    inventory_sheet = wb['3. Page Inventory']
    inventory_headers = [c.value for c in inventory_sheet[1]]
    inventory_rows = list(inventory_sheet.iter_rows(values_only=True))[1:]

    core_pages = []
    categories_list = []
    products_list = []

    cat_idx = 1
    prod_idx = 1

    def find_best_image(prod_name):
        pn = prod_name.lower()
        if "electropolished pipe" in pn or "electropolished tube" in pn or "sanitary electropolished" in pn:
            return "/electropolish_pipes.webp"
        elif "electropolished 3-a" in pn or "electropolished bs" in pn or "electropolished asme" in pn:
            return "/SS Dairy Fittings.webp"
        elif "weld neck flange" in pn:
            return "/Stainless Steel Flanges With Hub.webp"
        elif "flange" in pn:
            return "/Stainless Steel Flanges.webp"
        elif "elbow" in pn:
            return "/Stainless Steel Elbow.webp"
        elif "tee" in pn:
            return "/Stainless Steel Tee.webp"
        elif "reducer" in pn or "pipe cap" in pn or "stub end" in pn or "crosses" in pn or "socket weld fittings" in pn or "threaded forged fittings" in pn:
            return "/Stainless Steel Pipe Fittings.webp"
        elif "bend" in pn:
            return "/Stainless Steel Dairy Bend.webp"
        elif "welded (erw/efw) pipes" in pn:
            return "/flanges_pipes.webp"
        elif "duplex steel pipes" in pn or "super duplex pipes" in pn:
            return "/Duplex Pipe Fittings.webp"
        elif "304/304l round bar" in pn:
            return "/304 Stainless Steel Round Bar.webp"
        elif "316/316l round bar" in pn:
            return "/316 Stainless Steel Round Bar.webp"
        elif "duplex & super duplex round bar" in pn:
            return "/Duplex Steel Round Bar 3.webp"
        elif "17-4ph round bar" in pn:
            return "/17-4 PH Stainless Steel Round Bar.webp"
        elif "sms union" in pn or "tri-clover" in pn:
            return "/Stainless Steel SMS Union.webp"
        elif "bolt" in pn or "nuts" in pn or "stud" in pn or "fastener" in pn:
            return "/Stainless Steel Fasteners.webp"
        return "/flanges_pipes.webp"

    for row in inventory_rows:
        if not row[0]:
            continue
        page_type = row[0].strip()
        page_name = row[1].strip()
        url = row[2].strip()
        meta_title = row[3]
        meta_desc = row[5]
        h1 = row[7]
        robots = row[8]
        og_title = row[9]
        og_desc = row[10]
        og_image = row[11]
        priority = row[12]
        changefreq = row[13]
        breadcrumb = row[14]
        schema = row[15]
        
        slug = url.replace("https://steelmanufacturer.in/", "").strip("/")
        kw_info = keyword_map.get(url, {'Primary Keyword': '', 'Secondary Keywords': '', 'Long-tail Keywords': ''})
        
        if page_type == 'Core':
            core_pages.append({
                'Page Name': page_name,
                'URL': url,
                'Slug': slug,
                'Meta Title': meta_title,
                'Meta Description': meta_desc,
                'H1': h1,
                'Robots': robots,
                'OG Title': og_title,
                'OG Description': og_desc,
                'OG Image': og_image,
                'Priority': priority,
                'Changefreq': changefreq,
                'Breadcrumb': breadcrumb,
                'Schema': schema,
                **kw_info
            })
        elif page_type == 'Category':
            categories_list.append({
                'S.No': str(cat_idx),
                'Parent Category': page_name,
                'Category Slug': slug,
                'Category Meta Title': meta_title,
                'Category Meta Description': meta_desc,
                'Category H1': h1,
                'Robots': robots,
                'Sitemap priority': str(priority),
                'Changefreq': changefreq,
                'Breadcrumb trail': breadcrumb,
                'Schema on page': schema,
                'Primary KW': kw_info['Primary Keyword'],
                'Secondary KWs': kw_info['Secondary Keywords'],
                'Key Sub-Products to Cross-Link': ''
            })
            cat_idx += 1
        elif page_type == 'Product':
            content_info = product_content_map.get(url, {
                'Product Name': page_name,
                'Priority': 'Medium',
                'TL;DR': '',
                'H2 Outline': '',
                'Opening Copy': '',
                'Specs Table Raw': '',
                'Internal Links Raw': '',
                'Image Alt Texts': '',
                'Word Count Target': ''
            })
            
            category = 'Other'
            if breadcrumb:
                parts = [p.strip() for p in breadcrumb.split(">")]
                if len(parts) >= 3:
                    category = parts[2]
                    
            faqs = faq_map.get(url, [])
            
            products_list.append({
                'S.No': str(prod_idx),
                'Product Name': page_name,
                'Category': category,
                'URL Slug': slug,
                'Meta Title': meta_title,
                'Meta Description': meta_desc,
                'H1 Tag': h1,
                'Robots': robots,
                'Sitemap priority': str(priority),
                'Changefreq': changefreq,
                'Breadcrumb trail': breadcrumb,
                'Schema on page': schema,
                'Primary Keyword': kw_info['Primary Keyword'],
                'Secondary Keywords (3-5)': kw_info['Secondary Keywords'],
                'Product Description': content_info['Opening Copy'] or meta_desc,
                'TL;DR': content_info['TL;DR'],
                'H2 Outline': content_info['H2 Outline'],
                'Key Specs': content_info['Specs Table Raw'] or '',
                'Internal Links Raw': content_info['Internal Links Raw'] or '',
                'Image': find_best_image(page_name),
                'Image Alt Texts': content_info['Image Alt Texts'],
                'FAQs': faqs,
                **kw_info
            })
            prod_idx += 1

    # Populate category cross links dynamically
    for cat in categories_list:
        cat_prods = [p['Product Name'] for p in products_list if p['Category'] == cat['Parent Category']]
        cat['Key Sub-Products to Cross-Link'] = " | ".join(cat_prods)

    # Save output directly to src/data/
    os.makedirs(r"c:\Users\amity\OneDrive\Desktop\saksi-forge\src\data", exist_ok=True)
    
    with open(r"c:\Users\amity\OneDrive\Desktop\saksi-forge\src\data\categories.json", "w", encoding="utf-8") as f:
        json.dump(categories_list, f, indent=2, ensure_ascii=False)

    # Merge with original 170 products
    merged_products = []
    original_path = r"c:\Users\amity\OneDrive\Desktop\saksi-forge\original_products.json"
    if os.path.exists(original_path):
        with open(original_path, "r", encoding="utf-8") as f:
            original_list = json.load(f)
            
        new_slugs = {p['URL Slug'].lower().strip() for p in products_list}
        new_names = {p['Product Name'].lower().strip() for p in products_list}
        
        for orig_prod in original_list:
            orig_slug = orig_prod.get('URL Slug', '').lower().strip()
            orig_name = orig_prod.get('Product Name', '').lower().strip()
            if orig_slug in new_slugs or orig_name in new_names:
                continue
            merged_products.append(orig_prod)
        merged_products.extend(products_list)
    else:
        merged_products = products_list

    with open(r"c:\Users\amity\OneDrive\Desktop\saksi-forge\src\data\products.json", "w", encoding="utf-8") as f:
        json.dump(merged_products, f, indent=2, ensure_ascii=False)

    # Also save core_pages.json for the prerender engine to consume
    with open(r"c:\Users\amity\OneDrive\Desktop\saksi-forge\src\data\core_pages.json", "w", encoding="utf-8") as f:
        json.dump(core_pages, f, indent=2, ensure_ascii=False)

    print(f"Successfully compiled Excel data: {len(categories_list)} categories, {len(merged_products)} products, {len(core_pages)} core pages.")
    wb.close()

if __name__ == '__main__':
    compile_seo_data()
