import openpyxl
import json
import os

# Excel and data paths relative to the current script location or workspace root
workspace_dir = r"c:\Users\amity\OneDrive\Desktop\saksi-forge"
master_path = os.path.join(workspace_dir, "Sakshi_Forge_Full_SEO_Product_Pack_v2.xlsx")
batch3_path = os.path.join(workspace_dir, "Batch_3_Products_51_to_75_SEO_Content.xlsx")
batch4_path = os.path.join(workspace_dir, "Batch_4_Products_76_to_100_SEO_Content.xlsx")
batch5_path = os.path.join(workspace_dir, "Batch_5_Products_101_to_145_SEO_Content.xlsx")
products_json_path = os.path.join(workspace_dir, "frontend", "src", "data", "products.json")

missing_slugs = {
    'hastelloy-c276-flange': batch3_path,
    'hardox-400-plate': batch4_path,
    'stainless-steel-decorative-sheets': batch4_path,
    'stainless-steel-decorative-sheet': batch5_path,
    'stainless-steel-gold-sheet': batch5_path,
    'ss-rose-gold-sheet': batch5_path,
    'blue-decorative-sheets': batch5_path,
    'ss-linen-sheet': batch5_path,
    'water-ripple-sheet': batch5_path,
    'ss-electropolish-pipe': batch5_path,
    '304-stainless-steel-electropolish-pipe': batch5_path,
    'stainless-steel-honed-pipe': batch5_path
}

def find_best_image(prod_name):
    pn = prod_name.lower()
    if "electropolished pipe" in pn or "electropolished tube" in pn or "sanitary electropolished" in pn or "electropolish pipe" in pn or "honed pipe" in pn:
        return "/electropolish_pipes.webp"
    elif "electropolished 3-a" in pn or "electropolished bs" in pn or "electropolished asme" in pn:
        return "/SS Dairy Fittings.webp"
    elif "weld neck flange" in pn:
        return "/Stainless Steel Flanges With Hub.webp"
    elif "flange" in pn:
        return "/Stainless Steel Flanges.webp"
    elif "elbow" in pn:
        return "/Stainless Steel Elbow.webp"
    elif "tee" in pn:
        return "/Stainless Steel Tee.webp"
    elif "reducer" in pn or "pipe cap" in pn or "stub end" in pn or "crosses" in pn or "socket weld fittings" in pn or "threaded forged fittings" in pn:
        return "/Stainless Steel Pipe Fittings.webp"
    elif "bend" in pn:
        return "/Stainless Steel Dairy Bend.webp"
    elif "welded (erw/efw) pipes" in pn:
        return "/flanges_pipes.webp"
    elif "duplex steel pipes" in pn or "super duplex pipes" in pn:
        return "/Duplex Pipe Fittings.webp"
    elif "304/304l round bar" in pn:
        return "/304 Stainless Steel Round Bar.webp"
    elif "316/316l round bar" in pn:
        return "/316 Stainless Steel Round Bar.webp"
    elif "duplex & super duplex round bar" in pn:
        return "/Duplex Steel Round Bar 3.webp"
    elif "17-4ph round bar" in pn:
        return "/17-4 PH Stainless Steel Round Bar.webp"
    elif "sms union" in pn or "tri-clover" in pn:
        return "/Stainless Steel SMS Union.webp"
    elif "bolt" in pn or "nuts" in pn or "stud" in pn or "fastener" in pn:
        return "/Stainless Steel Fasteners.webp"
    elif "plate" in pn or "sheet" in pn:
        return "/shims.webp"
    return "/flanges_pipes.webp"

# 1. Load products.json
with open(products_json_path, "r", encoding="utf-8") as f:
    products = json.load(f)

# Find max S.No in products.json to continue sequence
max_sno = 0
for p in products:
    try:
        sno = int(p.get("S.No", 0))
        if sno > max_sno:
            max_sno = sno
    except ValueError:
        pass

print(f"Current products count: {len(products)}")
print(f"Max S.No: {max_sno}")

# 2. Load Master Product List
wb_master = openpyxl.load_workbook(master_path, data_only=True)
sheet_master = wb_master['Master Product List']
master_headers = [c for c in next(sheet_master.iter_rows(values_only=True))]
master_rows = list(sheet_master.iter_rows(values_only=True))[1:]

# 3. For each missing slug, find it in master list and batch file
added_count = 0
for slug, batch_path in missing_slugs.items():
    # Find in master sheet (exact first)
    master_row = None
    for row in master_rows:
        row_slug = str(row[3]).strip().lower() if row[3] else ""
        if slug == row_slug:
            master_row = row
            break
            
    if not master_row:
        # Fallback contains match
        for row in master_rows:
            row_slug = str(row[3]).strip().lower() if row[3] else ""
            if (len(slug) > 5 and slug in row_slug) or (len(row_slug) > 5 and row_slug in slug):
                master_row = row
                break
    
    if not master_row:
        print(f"Error: could not find slug '{slug}' in Master Product List.")
        continue
        
    master_data = dict(zip(master_headers, master_row))
    
    # Load details from batch sheet
    wb_batch = openpyxl.load_workbook(batch_path, data_only=True)
    sheet_batch = wb_batch.active
    batch_headers = [c for c in next(sheet_batch.iter_rows(values_only=True))]
    
    batch_row = None
    for row in sheet_batch.iter_rows(values_only=True):
        row_slug = str(row[1]).strip().lower() if row[1] else ""
        if slug == row_slug:
            batch_row = row
            break
            
    if not batch_row:
        for row in sheet_batch.iter_rows(values_only=True):
            row_slug = str(row[1]).strip().lower() if row[1] else ""
            if (len(slug) > 5 and slug in row_slug) or (len(row_slug) > 5 and row_slug in slug):
                batch_row = row
                break
            
    if not batch_row:
        # Fallback search by Product Name
        prod_name_master = str(master_data.get("Product Name")).strip().lower()
        for row in sheet_batch.iter_rows(values_only=True):
            row_name = str(row[0]).strip().lower() if row[0] else ""
            if prod_name_master == row_name:
                batch_row = row
                break
                
    if not batch_row:
        print(f"Error: could not find slug '{slug}' in batch file '{os.path.basename(batch_path)}'")
        wb_batch.close()
        continue
        
    batch_data = dict(zip(batch_headers, batch_row))
    wb_batch.close()
    
    # 4. Construct the product object
    max_sno += 1
    category = master_data.get("Category", "Other")
    prod_name = master_data.get("Product Name")
    
    # Breadcrumb trail classification
    parent_category = "Other"
    cat_lower = category.lower()
    if "flange" in cat_lower:
        parent_category = "Flanges"
    elif "sheet" in cat_lower:
        parent_category = "Sheets, Plates & Coils"
    elif "electropolish" in cat_lower:
        parent_category = "Electropolished Pipes"
        
    breadcrumb = f"Home > Products > {parent_category} > {prod_name}"
    
    # Generic FAQ bank for the product
    faqs = [
        {
            "q": f"What is the price of {prod_name} in India?",
            "a": f"The pricing of {prod_name} depends on the specifications, grade, size, and order volume. Sakshi Forge quotes within 30 minutes with a validity window - contact us with your requirements for today's best price."
        },
        {
            "q": f"Does Sakshi Forge provide Material Test Certificates (MTC) for {prod_name}?",
            "a": f"Yes, all our products are supplied with Material Test Certificates (MTC) as per EN 10204 3.1 along with relevant testing reports."
        },
        {
            "q": f"Which grades of {prod_name} are available?",
            "a": f"We manufacture and stock {prod_name} in various grades matching international standards. Please refer to the specifications or contact our team for availability."
        }
    ]
    
    new_prod = {
        "S.No": str(max_sno),
        "Product Name": prod_name,
        "Category": category,
        "URL Slug": master_data.get("URL Slug"),
        "Meta Title": batch_data.get("Meta Title") or master_data.get("Meta Title (<=60 chars)"),
        "Meta Description": batch_data.get("Meta Description") or master_data.get("Meta Description (<=160 chars)"),
        "H1 Tag": batch_data.get("H1") or master_data.get("H1 Tag"),
        "Robots": "index, follow",
        "Sitemap priority": "0.8",
        "Changefreq": "monthly",
        "Breadcrumb trail": breadcrumb,
        "Schema on page": "Product + BreadcrumbList + FAQPage",
        "Primary Keyword": master_data.get("Primary Keyword") or (prod_name + " manufacturer"),
        "Secondary Keywords (3-5)": master_data.get("Secondary Keywords (3-5)") or batch_data.get("Key Features"),
        "Product Description": batch_data.get("Product Description") or master_data.get("Meta Description (<=160 chars)"),
        "TL;DR": f"{prod_name} manufactured by Sakshi Forge, Mumbai. Premium quality materials, conforming to international standards. Get a quote in 30 minutes.",
        "H2 Outline": "1. Overview\n2. Specifications\n3. Key Features\n4. Applications\n5. Quality Assurance\n6. FAQs",
        "Key Specs": master_data.get("Key Specs") or f"Product | {prod_name}\nOrigin | Manufactured in Mumbai, India",
        "Internal Links Raw": master_data.get("Internal Links") or "",
        "Image": find_best_image(prod_name),
        "Image Alt Texts": f"{prod_name} manufacturer supplier India",
        "FAQs": faqs,
        "Secondary Keywords": master_data.get("Secondary Keywords (3-5)") or "",
        "Long-tail Keywords": f"what is {prod_name.lower()}; buy {prod_name.lower()} in India; {prod_name.lower()} price"
    }
    
    products.append(new_prod)
    added_count += 1
    print(f"Added product: {prod_name}")

wb_master.close()

# 5. Write back to products.json
with open(products_json_path, "w", encoding="utf-8") as f:
    json.dump(products, f, indent=2, ensure_ascii=False)

print(f"\nSuccessfully added {added_count} products. New total count: {len(products)}")
